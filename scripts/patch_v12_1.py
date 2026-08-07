#!/usr/bin/env python3
"""patch_v12_1.py (v2) — ERROR #58 cluster merge: v12 -> v12.1 (S1 window, data-DEC).

Merges the two overlapping clusters 'Pizzutilo et al (2020)' and
'Caragnano et al (2020)' (numerically identical printed descriptives =
same sample; ERROR #58, author ruling F2 (i), 2026-08-06, result-blind)
into one cluster_id 'CLUSTER Pizzutilo/Caragnano', following the
established merged-cluster label pattern ('CLUSTER Sandra/Ofogbe').

v2 scope (supersedes the in-session v1, which patched the data sheet
only and was never delivered or committed): cluster_id lives at THREE
sites in the self-contained-file architecture —
    data!cluster_id        (row level, 20 + 5 = 25 cells)
    provenance!cluster_id  (row level, 20 + 5 = 25 cells)
    lookup!cluster_id      (study level,  1 + 1 =  2 cells)
Total: 52 cells. The lookup/provenance mirror of merged clusters is
evidenced by the Sandra/Ofogbe precedent (CLUSTER label present at all
three sites). The 'study' column is untouched everywhere: both studies
remain distinct at the study level (120 studies); the three-level
nesting cluster_id/study/esid carries the merged cluster exactly as in
the precedent. Corpus cluster count 119 -> 118. (The estimation-set
count 114 -> 113 is asserted downstream in R/01_core.R, Commit B — a
derived subset, not a workbook cell.)

Convention: DEC-042 erratum pattern — transparent, script-based,
cell-exact, versioned; v12 remains archived, v12.1 supersedes it as the
sole analysis input; scripts/final_integrity_audit.py (v2) re-runs on
the output as the mandatory gate before freeze. Container-side
LibreOffice recalculation (established ritual, ERROR #9) restores the
formula caches openpyxl drops on save; it runs after this script and
before the audit.

Input : data/CER-COD_data_v12.xlsx    (canonical v12, DEC-042)
Output: data/CER-COD_data_v12_1.xlsx  (v12.1 erratum)

Self-verification: after writing, the script reloads both workbooks
(data_only=False) and performs a full cell-level diff over all sheets;
it aborts unless the diff set is exactly the 52 expected cells with the
expected old -> new content. Incumbent-asserts abort on any pre-state
deviation. Post-asserts re-derive the census per sheet and check the
three-site cross-consistency (data<->provenance outcome-keyed;
data<->lookup study-mapped).
"""

import hashlib
import sys

from openpyxl import load_workbook

# ---------------------------------------------------------------- constants
INPUT = "data/CER-COD_data_v12.xlsx"
OUTPUT = "data/CER-COD_data_v12_1.xlsx"

OLD_A = "Pizzutilo et al (2020)"
OLD_B = "Caragnano et al (2020)"
NEW = "CLUSTER Pizzutilo/Caragnano"
PRECEDENT = "CLUSTER Sandra/Ofogbe"

# (sheet, key column, expected count OLD_A, expected count OLD_B,
#  expected precedent count)
SITES = [
    ("data", "cluster_id", 20, 5, 6),
    ("provenance", "cluster_id", 20, 5, 6),
    ("lookup", "cluster_id", 1, 1, 2),
]

EXPECT_ROWS_DATA = 2852
EXPECT_ROWS_LOOKUP = 120
EXPECT_CLUSTERS_PRE = 119
EXPECT_CLUSTERS_POST = 118
EXPECT_STUDIES = 120
EXPECT_EDITS = 52


def die(msg: str) -> None:
    print(f"ABORT: {msg}")
    sys.exit(1)


def md5(path: str) -> str:
    h = hashlib.md5()
    with open(path, "rb") as fh:
        for chunk in iter(lambda: fh.read(1 << 20), b""):
            h.update(chunk)
    return h.hexdigest()


def col_index(ws, name: str) -> int:
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    hits = [i for i, v in enumerate(header, start=1) if v == name]
    if len(hits) != 1:
        die(f"{ws.title}: header '{name}' found {len(hits)}x (expected 1)")
    return hits[0]


def column_cells(ws, cidx: int):
    return [row[0] for row in ws.iter_rows(min_row=2, min_col=cidx,
                                           max_col=cidx)]


# ------------------------------------------------------------ 1. pre-asserts
print(f"[1] input : {INPUT}")
print(f"    md5   : {md5(INPUT)}")

wb = load_workbook(INPUT, data_only=False)
targets = {}   # sheet -> list of (row, cell, old_value)

for sheet, keycol, n_a, n_b, n_prec in SITES:
    if sheet not in wb.sheetnames:
        die(f"sheet '{sheet}' missing")
    ws = wb[sheet]
    cidx = col_index(ws, keycol)
    cells = column_cells(ws, cidx)
    vals = [c.value for c in cells if c.value is not None]

    rows_a = [c for c in cells if c.value == OLD_A]
    rows_b = [c for c in cells if c.value == OLD_B]
    rows_new = [c for c in cells if c.value == NEW]
    n_prec_found = sum(1 for v in vals if v == PRECEDENT)

    if len(rows_a) != n_a:
        die(f"{sheet}: incumbent '{OLD_A}' = {len(rows_a)} != {n_a}")
    if len(rows_b) != n_b:
        die(f"{sheet}: incumbent '{OLD_B}' = {len(rows_b)} != {n_b}")
    if rows_new:
        die(f"{sheet}: target label already present ({len(rows_new)}x)")
    if n_prec_found != n_prec:
        die(f"{sheet}: precedent '{PRECEDENT}' = {n_prec_found} != {n_prec}")

    uniq = len(set(vals))
    if uniq != EXPECT_CLUSTERS_PRE:
        die(f"{sheet}: unique cluster_id pre {uniq} != {EXPECT_CLUSTERS_PRE}")

    targets[sheet] = [(c.row, c, c.value) for c in rows_a + rows_b]
    print(f"[2] {sheet}: incumbent-asserts PASS "
          f"({n_a}+{n_b} targets, {uniq} clusters, precedent x{n_prec})")

# row censuses
ws_d = wb["data"]
n_data = sum(1 for c in column_cells(ws_d, col_index(ws_d, "study"))
             if c.value is not None)
if n_data != EXPECT_ROWS_DATA:
    die(f"data row census {n_data} != {EXPECT_ROWS_DATA}")
ws_l = wb["lookup"]
n_lk = sum(1 for c in column_cells(ws_l, col_index(ws_l, "study"))
           if c.value is not None)
if n_lk != EXPECT_ROWS_LOOKUP:
    die(f"lookup row census {n_lk} != {EXPECT_ROWS_LOOKUP}")

# ---------------------------------------------------------------- 2. patch
edited = []   # (sheet, coordinate, old)
for sheet, keycol, *_ in SITES:
    for r, cell, old in targets[sheet]:
        edited.append((sheet, cell.coordinate, old))
        cell.value = NEW

if len(edited) != EXPECT_EDITS:
    die(f"edited {len(edited)} cells != {EXPECT_EDITS}")

wb.save(OUTPUT)
print(f"[3] patched {len(edited)} cells -> '{NEW}'; wrote {OUTPUT}")
for sheet, *_ in SITES:
    rows = sorted(int(coord.lstrip("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))
                  for s, coord, _ in edited if s == sheet)
    print(f"    {sheet}: rows {rows}")

# ------------------------------------------------- 3. full cell-level diff
orig = load_workbook(INPUT, data_only=False)
new = load_workbook(OUTPUT, data_only=False)

if orig.sheetnames != new.sheetnames:
    die(f"sheet list changed: {orig.sheetnames} vs {new.sheetnames}")

diffs = []
for name in orig.sheetnames:
    so, sn = orig[name], new[name]
    max_r = max(so.max_row, sn.max_row)
    max_c = max(so.max_column, sn.max_column)
    for row_o, row_n in zip(
        so.iter_rows(min_row=1, max_row=max_r, max_col=max_c),
        sn.iter_rows(min_row=1, max_row=max_r, max_col=max_c),
    ):
        for co, cn in zip(row_o, row_n):
            if co.value != cn.value:
                diffs.append((name, co.coordinate, co.value, cn.value))

expected = {(s, coord) for s, coord, _ in edited}
got = {(s, coord) for s, coord, _, _ in diffs}
if got != expected:
    die(f"diff mismatch — unexpected: {sorted(got - expected)[:10]} "
        f"missing: {sorted(expected - got)[:10]} (total {len(diffs)})")
for s, coord, old, newv in diffs:
    if old not in (OLD_A, OLD_B) or newv != NEW:
        die(f"diff content mismatch at {s}!{coord}: {old!r} -> {newv!r}")

print(f"[4] full-workbook diff PASS: exactly {len(diffs)} cells differ "
      f"(data 25 / provenance 25 / lookup 2), all with expected content")

# ------------------------------------------------------------ 4. post-asserts
def colvals(wbx, sheet, name):
    ws = wbx[sheet]
    cidx = col_index(ws, name)
    return [c.value for c in column_cells(ws, cidx) if c.value is not None]

checks = []
for sheet, keycol, n_a, n_b, _ in SITES:
    vals = colvals(new, sheet, keycol)
    checks += [
        (f"{sheet} unique cluster_id", len(set(vals)), EXPECT_CLUSTERS_POST),
        (f"{sheet} merged-label count", vals.count(NEW), n_a + n_b),
        (f"{sheet} residue A", vals.count(OLD_A), 0),
        (f"{sheet} residue B", vals.count(OLD_B), 0),
    ]
checks += [
    ("data rows", len(colvals(new, "data", "cluster_id")), EXPECT_ROWS_DATA),
    ("provenance rows", len(colvals(new, "provenance", "cluster_id")),
     EXPECT_ROWS_DATA),
    ("lookup rows", len(colvals(new, "lookup", "cluster_id")),
     EXPECT_ROWS_LOOKUP),
    ("data unique study", len(set(colvals(new, "data", "study"))),
     EXPECT_STUDIES),
]

# three-site cross-consistency
ws_d, ws_p, ws_l = new["data"], new["provenance"], new["lookup"]
d_out = [c.value for c in column_cells(ws_d, col_index(ws_d, "outcome"))]
d_cl = [c.value for c in column_cells(ws_d, col_index(ws_d, "cluster_id"))]
d_st = [c.value for c in column_cells(ws_d, col_index(ws_d, "study"))]
p_out = [c.value for c in column_cells(ws_p, col_index(ws_p, "outcome"))]
p_cl = [c.value for c in column_cells(ws_p, col_index(ws_p, "cluster_id"))]
l_st = [c.value for c in column_cells(ws_l, col_index(ws_l, "study"))]
l_cl = [c.value for c in column_cells(ws_l, col_index(ws_l, "cluster_id"))]

pmap = dict(zip(p_out, p_cl))
mism_dp = sum(1 for o, c in zip(d_out, d_cl)
              if o is not None and pmap.get(o) != c)
lmap = dict(zip(l_st, l_cl))
mism_dl = sum(1 for s, c in zip(d_st, d_cl)
              if s is not None and lmap.get(s) != c)
checks += [
    ("data<->provenance cluster mismatches (outcome key)", mism_dp, 0),
    ("data<->lookup cluster mismatches (study map)", mism_dl, 0),
]

fail = [(k, g, e) for k, g, e in checks if g != e]
if fail:
    die(f"post-asserts FAILED: {fail}")
print("[5] post-asserts PASS:")
for k, g, _ in checks:
    print(f"      {k} = {g}")

print(f"[6] output: {OUTPUT}")
print(f"    md5 (pre-recalc; container recalc + audit gate follow): "
      f"{md5(OUTPUT)}")
print("DONE")
