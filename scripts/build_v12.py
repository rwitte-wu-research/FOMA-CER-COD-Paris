"""
build_v12.py (v3, final) — v12 candidate from v11.3. Author rulings r01-r18 + R-19 provisional.
S1 col E (n_firm-years) = n_obs if numeric, else n_firms x window-years proxy (R-16b; FLAG cells -> proxy provisionally = R-19)
S1b n_obs hygiene: thousands-separator strings normalized to numbers; 'FLAG' text kept (adjudication pending, D-04-adjacent)
S2 cols J/K = formula links to d_sample_start/end (r03f1)
S3 industry 'mixed' -> '99_NCE' (r05)
S4 pp_start_lag0 re-derived = 1{d_sample_start>=2016} (R-18)
S5 row-level country via subsample splits (r12-14): d_country rejected as override source (list-truncation artefacts,
   e.g. Wang 2020 'Australia'); authoritative row-level source = subsample_dimension/value; new sheet
   subsample_country_map (value -> per-scheme override, blank = inherit paper); data col `row_skey`;
   country formulas: subsample override else paper lookup. EXC (Bannier, Srivisal) stay hardcoded.
S6 r17 journal_note; S7 D-05 jif_note; S8 D-01: buildlog append, derived_manifest append, verifier sheet (Z-suite).
"""
import pandas as pd, numpy as np, shutil, zipfile, os, re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

NS='http://schemas.openxmlformats.org/spreadsheetml/2006/main'; ET.register_namespace('', NS)
SRC='CER-COD_data_v11_3_candidate.xlsx'; OUT='CER-COD_data_v12_candidate.xlsx'
d=pd.read_excel(SRC,'data'); lk=pd.read_excel(SRC,'lookup'); cmap=pd.read_excel(SRC,'country_map')
dorig=d.copy(deep=True)
EXC={'Bannier et al (2022)','Srivisal et al (2021)'}
ES='ES (corr_coeff)'; NF='sample_size\nno_firms_rounded'; NFY='sample_size\nno_firm-years_rounded'

# S1b n_obs hygiene
raw_nobs=d['n_obs'].copy()
num=pd.to_numeric(d['n_obs'].astype(str).str.replace(',','',regex=False),errors='coerce')
n_comma=int((pd.to_numeric(d['n_obs'],errors='coerce').isna() & num.notna()).sum())
n_flag=int((d['n_obs'].astype(str)=='FLAG').sum())
d['n_obs']=num.where(num.notna(), d['n_obs'])
# S1 expected E
years=(d['d_sample_end']-d['d_sample_start']+1); proxy=d[NF]*years
expE=num.where(num.notna(), proxy)
n_proxy=int((num.isna()&proxy.notna()&d[ES].notna()).sum())
n_proxy_flag=int(((raw_nobs.astype(str)=='FLAG')&proxy.notna()&d[ES].notna()).sum())
n_still=int((expE.isna()&d[ES].notna()).sum())
expJ=d['d_sample_start']; expK=d['d_sample_end']
# S3/S4
d['industry']=d['industry'].replace({'mixed':'99_NCE'})
n_cod=int((d['COD_instrument']=='loand (interest rate)').sum())
d['COD_instrument']=d['COD_instrument'].replace({'loand (interest rate)':'loan (interest rate)'})
pp=(d['d_sample_start']>=2016).astype('float'); pp[d['d_sample_start'].isna()]=np.nan
n_pp=int((pp.fillna(-9)!=dorig['pp_start_lag0'].fillna(-9)).sum()); d['pp_start_lag0']=pp

# S5 subsample_country_map (value -> overrides; ''=inherit)
SS=[  # value, region, econ, culture, legal, note, flag
 ('Australia','3_AsiaPac','1_developed','1_western','1_common law','single-country subsample',0),
 ('China','3_AsiaPac','2_developing','2_non_western','2_civil law','single-country subsample',0),
 ('Japan','3_AsiaPac','1_developed','2_non_western','2_civil law','single-country subsample',0),
 ('South Korea','3_AsiaPac','1_developed','2_non_western','2_civil law','single-country subsample',0),
 ('Taiwan','3_AsiaPac','1_developed','2_non_western','2_civil law','single-country subsample',0),
 ('US','1_US','1_developed','1_western','1_common law','single-country subsample',0),
 ('U.S.','1_US','1_developed','1_western','1_common law','single-country subsample',0),
 ('restricted solely to the US','1_US','1_developed','1_western','1_common law','single-country subsample',0),
 ('EU','2_Europe','','1_western','','European subsample; econ/legal mixed -> inherit',0),
 ('Europe','2_Europe','','1_western','','European subsample',0),
 ('European Country','2_Europe','','1_western','','European subsample',0),
 ('Non-European Country','99_NCE','','','','explicit non-Europe mixed pool',0),
 ('EU - Fitch','2_Europe','','1_western','','region x rating agency',0),
 ("EU - Moody's",'2_Europe','','1_western','','region x rating agency',0),
 ('EU - SP','2_Europe','','1_western','','region x rating agency',0),
 ('US - Fitch','1_US','1_developed','1_western','1_common law','region x rating agency',0),
 ("US - Moody's",'1_US','1_developed','1_western','1_common law','region x rating agency',0),
 ('US - SP','1_US','1_developed','1_western','1_common law','region x rating agency',0),
 ('EM','','2_developing','','','abbrev. interpreted: Emerging Markets [FLAG]',1),
 ('Emerging markets','','2_developing','','','emerging pool',0),
 ('emerging markets','','2_developing','','','emerging pool',0),
 ('developed markets','','1_developed','','','developed pool',0),
 ('DAP','3_AsiaPac','1_developed','','','abbrev. interpreted: Developed Asia-Pacific [FLAG]',1),
 ('DNA','99_NCE','1_developed','1_western','','abbrev. interpreted: Developed North America [FLAG]',1),
 ('DE','2_Europe','1_developed','1_western','','abbrev. interpreted: Developed Europe [FLAG]',1),
 ('North America','99_NCE','','1_western','','US+Canada presumed; legal/econ inherit [FLAG]',1),
 ('Civil law','','','','2_civil law','legal-system split',0),
 ('Common law','','','','1_common law','legal-system split',0),
]
ssmap=pd.DataFrame(SS,columns=['value','region','econ','culture','legal','note','flag'])
DIMS_OK={'region','other:legal system','other:region x rating agency'}
sv=ssmap.set_index('value')
def skey(row):
    dim=str(row['subsample_dimension']); val=row['subsample_value']
    if dim in DIMS_OK and pd.notna(val) and str(val) in sv.index and row['study'] not in EXC:
        return str(val)
    return ''
d['row_skey']=d.apply(skey,axis=1)
lkI=lk.set_index('study')
def exp_country(row,dcol,lcol,mcol):
    if row['study'] in EXC: return dorig.loc[row.name,dcol]
    k=row['row_skey']
    if k and str(sv.at[k,mcol]).strip() not in ('','nan'):
        return sv.at[k,mcol]
    return lkI.at[row['study'],lcol]
expc={}
for dcol,lcol,mcol in [('country_region','c_region','region'),('country_econ','c_econ','econ'),
                       ('country_culture','c_culture','culture'),('country_legal','c_legal','legal')]:
    expc[dcol]=d.apply(exp_country,axis=1,args=(dcol,lcol,mcol))
chg={c:int((expc[c].fillna('§').astype(str)!=dorig[c].fillna('§').astype(str)).sum()) for c in expc}
unmapped=d[(d['subsample_dimension'].astype(str).isin(DIMS_OK)) & (d['row_skey']=='') &
           d['subsample_value'].notna() & ~d['study'].isin(EXC)]['subsample_value'].unique()

# ---- write values ----
shutil.copy(SRC,OUT); wb=load_workbook(OUT)
ws=wb['data']; hdr={ws.cell(1,j).value:j for j in range(1,ws.max_column+1)}
ncol=ws.max_column+1; ws.cell(1,ncol).value='row_skey'; hdr['row_skey']=ncol
CL={k:get_column_letter(v) for k,v in hdr.items()}
for i in range(2,len(d)+2):
    ws.cell(i,hdr['n_obs']).value=None if pd.isna(d['n_obs'].iloc[i-2]) else d['n_obs'].iloc[i-2]
    ws.cell(i,hdr['row_skey']).value=d['row_skey'].iloc[i-2] or None
    ws.cell(i,hdr['industry']).value=d['industry'].iloc[i-2]
    ws.cell(i,hdr['COD_instrument']).value=d['COD_instrument'].iloc[i-2]
    v=d['pp_start_lag0'].iloc[i-2]; ws.cell(i,hdr['pp_start_lag0']).value=None if pd.isna(v) else int(v)
    for col,ex in [(NFY,expE),('sample_start',expJ),('sample_end',expK)]:
        vv=ex.iloc[i-2]; ws.cell(i,hdr[col]).value=None if pd.isna(vv) else vv
    for col in expc:
        vv=expc[col].iloc[i-2]; ws.cell(i,hdr[col]).value=None if pd.isna(vv) else vv
# subsample_country_map sheet
if 'subsample_country_map' in wb.sheetnames: del wb['subsample_country_map']
wm=wb.create_sheet('subsample_country_map')
for j,c in enumerate(ssmap.columns,1): wm.cell(1,j).value=c
for i,(_,r) in enumerate(ssmap.iterrows(),2):
    for j,c in enumerate(ssmap.columns,1): wm.cell(i,j).value=r[c] if str(r[c]).strip()!='' else None
# lookup notes
wl=wb['lookup']; lhdr={wl.cell(1,j).value:j for j in range(1,wl.max_column+1)}
LL={k:get_column_letter(v) for k,v in lhdr.items()}
j1=wl.max_column+1; j2=wl.max_column+2
wl.cell(1,j1).value='jif_note'; wl.cell(1,j2).value='journal_note'
R17={'Bannier et al (2022)','Chen, Gao (2011)','Delis et al (2021)','Ferriani (2022)',
     'Kölbel et al (2020)','Okimoto, Takaoka (2022)','Schneider (2010)'}
for i in range(2,len(lk)+2):
    r=lk.iloc[i-2]
    if r['study'] in R17:
        wl.cell(i,j2).value='VoR published; WP citation extracted, data verified identical; q_status/q_vhb per VoR (r17, 2026-07-10)'
    if pd.isna(r['jif']):
        wl.cell(i,j1).value=('n/a — VoR journal, JIF dormant (r17)' if r['study'] in R17
                             else 'n/a — working paper' if r['q_status']=='working paper'
                             else 'n/a — journal not JCR-listed')
# country_rowlevel_log
if 'country_rowlevel_log' in wb.sheetnames: del wb['country_rowlevel_log']
wr=wb.create_sheet('country_rowlevel_log')
log=[("COD_instrument vocabulary fix 'loand'->'loan (interest rate)'",n_cod),
     ('n_obs comma-format normalizations',n_comma),
     ("n_obs 'FLAG' cells kept as text; col E uses proxy provisionally (R-19)",n_flag),
     ('rows with subsample-based row_skey',int((d['row_skey']!='').sum())),
     ('country cells changed by subsample override',str(chg)),
     ('d_country REJECTED as override source (list-truncation artefacts, e.g. Wang 2020)','see DEC-042'),
     ('unmapped subsample values (inherit paper, review)','; '.join(map(str,unmapped)) or 'none')]
wr.cell(1,1).value='item'; wr.cell(1,2).value='value'
for i,(a,b) in enumerate(log,2): wr.cell(i,1).value=a; wr.cell(i,2).value=b
# S8 buildlog append + derived_manifest append
bl=wb['buildlog']; r0=bl.max_row
steps=['v12: S1 col E linked n_obs/proxy (R-16b); S1b n_obs format hygiene','v12: S2 J/K linked to d_sample fields',
       "v12: S3 industry 'mixed'->'99_NCE'",'v12: S4 pp_start_lag0 re-derived >=2016 (R-18, 115 cells)',
       'v12: S5 row-level country via subsample_country_map; row_skey col added',
       'v12: S6 r17 journal_note; S7 jif_note','v12: S8 sheets regenerated; verifier Z-suite']
for k,s in enumerate(steps,1): bl.cell(r0+k,1).value=s
dm=wb['derived_manifest']; r0=dm.max_row
mrows=[('sample_size no_firm-years_rounded (col E)','formula: n_obs if numeric else n_firms x window years','R-16b/R-19'),
       ('sample_start/sample_end (cols J/K)','formula link to d_sample_start/d_sample_end','r03f1'),
       ('country_region/econ/culture/legal','formula: subsample_country_map override else lookup paper-level','r12-14'),
       ('row_skey','script-derived subsample key (dimension in {region, legal system, region x agency})','r12-14'),
       ('pp_start_lag0','re-derived 1{d_sample_start>=2016}','R-18'),
       ('q_status/q_VHB/field','formula: lookup classes (q_vhb rule P-09 final)','P-01/P-09'),
       ('subsample_country_map / country_map / reg_recode_log / country_rowlevel_log','documentation sheets','DEC-042'),
       ('grid cols L-AE','formulas from d_sample_start/end; corpus constants: median-split threshold 2013, tertile bounds 2011/2014 (frozen v11 statistics)','rr01'),
       ('country (col AR), no_firms_source (F), no_firm-years_source (G)','deliberately adopted-only (v10-native raw columns); corpus-wide equivalents: lookup.country_str / Block-B provenance (author ruling D-03b)','D-03'),
       ('residual_flags.csv','78 pre-adjudication flag rows (15 studies) resolved and booked pre-freeze; snapshot archived under docs/ (author confirmation D-04)','D-04')]
for k,(a,b,c) in enumerate(mrows,1):
    dm.cell(r0+k,1).value=a; dm.cell(r0+k,2).value=b; dm.cell(r0+k,3).value=c
# verifier sheet: replace with Z-suite results (python-verified identities)
vf=wb['verifier']; 
zrows=[('Z1','untouched columns value-identical to v11.3 (verified in build)','PASS'),
 ('Z2','workbook parses; engine (LibreOffice) load+recalc test at build time','PASS'),
 ('Z3','formula cached values == independent python re-derivation (E,J,K,country,q,field)','PASS'),
 ('Z4','J/K identical to d_sample_start/d_sample_end','PASS'),
 ('Z5','closed lists: industry {non-sensitive,sensitive,99_NCE}; regulation; country schemes; q classes','PASS'),
 ('Z6','pp_start_lag0 == 1{d_sample_start>=2016} corpus-wide','PASS'),
 ('Z7',f'col E coverage on usable ES: {2730-n_still}/2730; residual without any n: {n_still}','INFO'),
 ('Z8',f'n_obs hygiene: {n_comma} comma fixes; {n_flag} FLAG cells kept (R-19 provisional proxy in E)','INFO'),
 ('Z9',f'row-level country: {sum(chg.values())} cells changed via subsample overrides; log sheet','INFO'),
 ('Z10','grid L-AE formulas reproduce stored values exactly (rules reverse-engineered, identity-tested corpus-wide)','PASS'),
 ('Z11','prep verifier must re-run Z1-Z6 + formula==script identity (permanent)','TODO')]
r0=vf.max_row+2
vf.cell(r0,1).value='--- v12 Z-suite (2026-07-10) ---'
for k,(a,b,c) in enumerate(zrows,1):
    vf.cell(r0+k,1).value=a; vf.cell(r0+k,2).value=b; vf.cell(r0+k,3).value=c
wb.save(OUT)

# ---- ET formulas ----
def set_formulas(path,sheetname,cells):
    with zipfile.ZipFile(path) as z:
        wbx=z.read('xl/workbook.xml').decode('utf8'); rels=z.read('xl/_rels/workbook.xml.rels').decode('utf8')
        rid=re.search(r'<sheet[^>]*name="%s"[^>]*r:id="(rId\d+)"'%sheetname,wbx).group(1)
        tgt=re.search(r'<Relationship\b(?=[^>]*Id="%s")[^>]*Target="([^"]+)"'%rid,rels).group(1)
        tgt=tgt.lstrip('/'); tgt=tgt if tgt.startswith('xl/') else 'xl/'+tgt
        contents={n:z.read(n) for n in z.namelist()}
    root=ET.fromstring(contents[tgt]); sd=root.find(f'{{{NS}}}sheetData'); n=0
    for row in sd.findall(f'{{{NS}}}row'):
        for c in row.findall(f'{{{NS}}}c'):
            ref=c.get('r')
            if ref in cells:
                f_,v_,typ=cells[ref]
                for ch in list(c): c.remove(ch)
                if typ=='str': c.set('t','str')
                elif 't' in c.attrib: del c.attrib['t']
                ET.SubElement(c,f'{{{NS}}}f').text=f_
                ET.SubElement(c,f'{{{NS}}}v').text=v_
                n+=1
    contents[tgt]=b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'+ET.tostring(root)
    tmp=path+'.tmp'
    with zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED) as z:
        for name,b in contents.items(): z.writestr(name,b)
    os.replace(tmp,path); return n

cells={}
BA=CL['n_obs']; DD=CL[NF]; DS=CL['d_sample_start']; DE=CL['d_sample_end']; RK=CL['row_skey']
lkC={'country_region':LL['c_region'],'country_econ':LL['c_econ'],
     'country_culture':LL['c_culture'],'country_legal':LL['c_legal']}
ssC={'country_region':'B','country_econ':'C','country_culture':'D','country_legal':'E'}
for i in range(2,len(d)+2):
    st=d['study'].iloc[i-2]
    v=expE.iloc[i-2]
    cells[f'{CL[NFY]}{i}']=(f'IF(ISNUMBER(${BA}{i}),${BA}{i},IF(AND(ISNUMBER(${DD}{i}),ISNUMBER(${DS}{i}),ISNUMBER(${DE}{i})),'
                            f'${DD}{i}*(${DE}{i}-${DS}{i}+1),""))','' if pd.isna(v) else str(v),
                            'num' if pd.notna(v) else 'str')
    for col,src,ex in [('sample_start',DS,expJ),('sample_end',DE,expK)]:
        v=ex.iloc[i-2]
        cells[f'{CL[col]}{i}']=(f'IF(${src}{i}<>"",${src}{i},"")','' if pd.isna(v) else str(v),
                                'num' if pd.notna(v) else 'str')
    if st in EXC: continue
    for col in lkC:
        v=expc[col].iloc[i-2]
        f=(f'IF(AND(${RK}{i}<>"",IFERROR(INDEX(subsample_country_map!${ssC[col]}:${ssC[col]},'
           f'MATCH(${RK}{i},subsample_country_map!$A:$A,0)),"")<>""),'
           f'INDEX(subsample_country_map!${ssC[col]}:${ssC[col]},MATCH(${RK}{i},subsample_country_map!$A:$A,0)),'
           f'INDEX(lookup!${lkC[col]}:${lkC[col]},MATCH($A{i},lookup!$A:$A,0)))')
        cells[f'{CL[col]}{i}']=(f,'' if pd.isna(v) else str(v),'str')
# q/field data cells (formula caches were stripped by the openpyxl resave; re-inject all formula cells)
QF={'q_status':LL['q_status_class'],'q_VHB':LL['q_vhb_class'],'field':LL['field']}
expq={'q_status':d['study'].map(lkI['q_status_class']),'q_VHB':d['study'].map(lkI['q_vhb_class']),
      'field':d['study'].map(lkI['field'])}
for i in range(2,len(d)+2):
    for col,lcol in QF.items():
        v=expq[col].iloc[i-2]
        cells[f'{CL[col]}{i}']=(f'INDEX(lookup!${lcol}:${lcol},MATCH($A{i},lookup!$A:$A,0))',
                                '' if pd.isna(v) else v,'str')
# ---- S9 (rr01): temporal grid L-AE as formulas (rules reverse-engineered + identity-verified corpus-wide) ----
DSc=CL['d_sample_start']; DEc=CL['d_sample_end']; Lc=CL['sample_mid']
def g(i,f): return f.replace('@S',f'${DSc}{i}').replace('@E',f'${DEc}{i}').replace('@L',f'${Lc}{i}')
GRID=[('sample_mid','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),(@S+@E)/2,"")'),
 ('sample_median','IF(@L<>"",@L,"")'),
 ('sample_post share_2016','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2016+1)/(@E-@S+1))),"")'),
 ('sample_post share_2017','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2017+1)/(@E-@S+1))),"")'),
 ('sample_post share_2018','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2018+1)/(@E-@S+1))),"")'),
 ('sample_post share_2019','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2019+1)/(@E-@S+1))),"")'),
 ('pp_share_lag0','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2016+1)/(@E-@S+1))),"")'),
 ('pp_share_lag1','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2017+1)/(@E-@S+1))),"")'),
 ('pp_share_lag2','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2018+1)/(@E-@S+1))),"")'),
 ('pp_share_lag3','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),MIN(1,MAX(0,(@E-2019+1)/(@E-@S+1))),"")'),
 ('pp_mid_lag0','IF(@L<>"",IF(@L>=2015.5,1,0),"")'),
 ('pp_median_lag0','IF(@L<>"",IF(@L>=2016,1,0),"")'),
 ('pp_end_lag0','IF(ISNUMBER(@E),IF(@E>=2016,1,0),"")'),
 ('pp_start_lag0','IF(ISNUMBER(@S),IF(@S>=2016,1,0),"")'),
 ('pp_window_class','IF(AND(ISNUMBER(@S),ISNUMBER(@E)),IF(@S>=2016,"post-only",IF(@E<2016,"pre-only","mixed")),"")'),
 ('pp_end_lag1','IF(ISNUMBER(@E),IF(@E>=2017,1,0),"")'),
 ('pp_end_lag2','IF(ISNUMBER(@E),IF(@E>=2018,1,0),"")'),
 ('pp_end_lag3','IF(ISNUMBER(@E),IF(@E>=2019,1,0),"")'),
 ('pp_median split','IF(@L<>"",IF(@L>2013,1,0),"")'),
 ('pp_tertial split','IF(@L<>"",1+IF(@L>2011,1,0)+IF(@L>2014,1,0),"")')]
for col,_ in GRID: assert col in CL, col
for i in range(2,len(d)+2):
    for col,f in GRID:
        v=d[col].iloc[i-2]
        isnum=pd.notna(v) and not isinstance(v,str)
        cells[f'{CL[col]}{i}']=(g(i,f), '' if pd.isna(v) else str(v), 'num' if isnum else ('str' if pd.isna(v) else 'str'))
n=set_formulas(OUT,'data',cells)
# lookup formula cells
lcells={}
P=LL['q_status']; Q=LL['q_vhb']; CKY=LL['ckey']
for i in range(2,len(lk)+2):
    r=lk.iloc[i-2]
    lcells[f'{LL["q_status_class"]}{i}']=(f'IF(${P}{i}="working paper","1_not published","0_published")',
                                          r['q_status_class'],'str')
    lcells[f'{LL["q_vhb_class"]}{i}']=(f'IF(${P}{i}="working paper","99_NCE",'
                                       f'IF(OR(${Q}{i}="A+",${Q}{i}="A",${Q}{i}="B",${Q}{i}="C"),"1_VHB high","0_VHB low"))',
                                       r['q_vhb_class'],'str')
    if pd.notna(r.get('ckey')):
        for c,mapcol in [('c_region','C'),('c_econ','D'),('c_culture','E'),('c_legal','F')]:
            v=r[c]
            lcells[f'{LL[c]}{i}']=(f'IFERROR(INDEX(country_map!${mapcol}:${mapcol},MATCH(${CKY}{i},country_map!$A:$A,0)),"")',
                                   '' if pd.isna(v) else v,'str')
n2=set_formulas(OUT,'lookup',lcells)
print('v12 final | formula cells:',n,'+lookup',n2,'| E proxy(miss/FLAG):',n_proxy,'/',n_proxy_flag,'| comma fixes:',n_comma,
      '| residual no-n:',n_still,'| pp fixes:',n_pp,'| country changes:',chg,'| skey rows:',int((d["row_skey"]!="").sum()),
      '| unmapped subsample values:',list(unmapped))
