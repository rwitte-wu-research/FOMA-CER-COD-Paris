"""
patch_v11_1.py — v11.1 candidate build (DEC-042 draft scope) — v2 (final)
Inputs : CER-COD_data_v11.xlsx (canonical, DEC-041) · CER-COD_data_v10.xlsx (legacy layers)
Outputs: CER-COD_data_v11_1_candidate.xlsx · verifier report (stdout)
Reads  : country_lookup_extension_v11_1.csv (rule table, committed alongside)
Scope  : S1 q-layer carry (data rows legacy-semantics; lookup = canonical raw layer incl. VHB letters)
         S2 country schemes region/culture/legal for new studies (rule-based extension, parse-homogeneous)
         S3 country_econ off-list recode emerging* -> 2_developing
         S4 regulation 'multi' -> '99_NCE'
         S5 verifier W1-W9 (content identity of untouched columns; closed lists; coverage)
Open   : q_VHB analysis operationalization (P-01) — new-study data rows deliberately NOT filled;
         raw letters live in lookup; prep derives the moderator per the P-01 ruling.
"""
import pandas as pd, numpy as np, shutil
from openpyxl import load_workbook

SRC11='/mnt/project/CERCOD_data_v11.xlsx'; SRC10='/mnt/project/CERCOD_data_v10.xlsx'
OUT='CER-COD_data_v11_1_candidate.xlsx'; EXT='country_lookup_extension_v11_1.csv'

d11=pd.read_excel(SRC11,'data'); lk11=pd.read_excel(SRC11,'lookup')
d10=pd.read_excel(SRC10,'data'); sl10=pd.read_excel(SRC10,'source_lookup')
orig=d11.copy(deep=True)

QCOLS=['q_status','q_VHB','field']
for c in QCOLS:
    assert (d10.groupby('study')[c].nunique(dropna=False)<=1).all(), f'{c} not study-unique in v10'
qmap=d10.groupby('study')[QCOLS].first()
legacy=set(d10['study'].unique()); mL=d11['study'].isin(legacy); mN=~mL

# S1a legacy rows: carry v10 study-level q values (legacy semantics, internally consistent)
for c in QCOLS:
    t=mL & d11[c].isna() & d11['study'].map(qmap[c]).notna()
    d11.loc[t,c]=d11.loc[t,'study'].map(qmap[c])

# S1b new-study rows: q_status from lookup, harmonized to legacy codes; q_VHB left open (P-01)
lkq=lk11.set_index('study')
smap={'published':'0_published','in press':'0_published','working paper':'1_not published'}  # in-press mapping = P-04
fill=d11.loc[mN,'study'].map(lkq['q_status']).map(smap)
t=mN & d11['q_status'].isna() & fill.reindex(d11.index).notna()
d11.loc[t,'q_status']=d11.loc[t,'study'].map(lkq['q_status']).map(smap)

# S1c lookup enrichment (canonical raw layer) for legacy studies
srcmap=d10.groupby('study')['source'].first()
sl10k=sl10.set_index(sl10['source'].astype(str).str.strip())
lk11=lk11.set_index('study')
for c in ['q_status','q_vhb','jif','journal_full','year_vor']: lk11[c]=lk11[c].astype(object)
statmap_rev={'0_published':'published','1_not published':'working paper'}
for st in lk11.index:
    if st not in qmap.index: continue
    if pd.isna(lk11.at[st,'q_status']) and pd.notna(qmap.at[st,'q_status']):
        lk11.at[st,'q_status']=statmap_rev.get(qmap.at[st,'q_status'],qmap.at[st,'q_status'])
    ref=str(srcmap.get(st,'')).strip()
    if ref in sl10k.index:
        row=sl10k.loc[ref]; row=row.iloc[0] if isinstance(row,pd.DataFrame) else row
        raw=str(row.get('journal vhb ranking','')).strip()
        if pd.isna(lk11.at[st,'q_vhb']) and raw in {'A+','A','B','C','D'}:
            lk11.at[st,'q_vhb']=raw                       # 'nicht gelistet'/'-' stays missing (DEC-040)
        if pd.isna(lk11.at[st,'jif']) and pd.notna(row.get('jounal Journal impact factor')):
            lk11.at[st,'jif']=row['jounal Journal impact factor']
        if pd.isna(lk11.at[st,'journal_full']) and pd.notna(row.get('journal ')):
            lk11.at[st,'journal_full']=row['journal ']
        if pd.isna(lk11.at[st,'year_vor']) and pd.notna(row.get('publication year')):
            lk11.at[st,'year_vor']=row['publication year']
lk11=lk11.reset_index()

# S2 country schemes for new studies via rule table (parse-homogeneous; 6 rules flagged for spot-check)
ext=pd.read_csv(EXT); E=ext.to_dict('records')
def classify(s):
    hits=[e for e in E if str(s).startswith(e['mod_country_key'])] or [e for e in E if e['mod_country_key'] in str(s)]
    return max(hits,key=lambda e:len(e['mod_country_key'])) if hits else None
cmap={st:classify(c) for st,c in lk11.set_index('study')['mod_country'].dropna().items()}
assert all(v is not None for v in cmap.values()), 'unmatched country strings'
for dcol,ecol in [('country_region','region'),('country_culture','culture'),('country_legal','legal')]:
    fill=d11.loc[mN,'study'].map({k:v[ecol] for k,v in cmap.items()})
    t=mN & d11[dcol].isna() & fill.reindex(d11.index).notna()
    d11.loc[t,dcol]=d11.loc[t,'study'].map({k:v[ecol] for k,v in cmap.items()})

# S3 / S4
d11.loc[d11['country_econ'].astype(str).str.startswith('emerging'),'country_econ']='2_developing'
for c in ['regulation_sample_start','regulation_sample_end']:
    d11.loc[d11[c]=='multi',c]='99_NCE'

# write
shutil.copy(SRC11,OUT); wb=load_workbook(OUT); ws=wb['data']
hdr={ws.cell(1,j).value:j for j in range(1,ws.max_column+1)}
TOUCH=['q_status','q_VHB','field','country_region','country_culture','country_legal',
       'country_econ','regulation_sample_start','regulation_sample_end']
for col in TOUCH:
    j=hdr[col]
    for i,v in enumerate(d11[col].tolist(),start=2):
        ws.cell(i,j).value=None if pd.isna(v) else v
wl=wb['lookup']; lhdr={wl.cell(1,j).value:j for j in range(1,wl.max_column+1)}
for col in ['q_status','q_vhb','jif','journal_full','year_vor']:
    j=lhdr[col]
    for i,v in enumerate(lk11[col].tolist(),start=2):
        wl.cell(i,j).value=None if pd.isna(v) else v
wb.save(OUT)

# S5 verifier
new=pd.read_excel(OUT,'data'); lknew=pd.read_excel(OUT,'lookup')
UNTOUCH=[c for c in orig.columns if c not in TOUCH]
ok=all(orig[c].fillna('§').astype(str).equals(new[c].fillna('§').astype(str)) for c in UNTOUCH)
print('W1 untouched-column content identity:', 'PASS' if ok else 'FAIL')
def cl(col,allowed):
    bad=set(new[col].dropna().unique())-set(allowed); return 'PASS' if not bad else f'FAIL {bad}'
print('W2 country_econ:',cl('country_econ',{'1_developed','2_developing','99_NCE'}))
print('W3 country_region:',cl('country_region',{'1_US','2_Europe','3_AsiaPac','99_NCE'}))
print('W4 country_culture:',cl('country_culture',{'1_western','2_non_western','99_NCE'}))
print('W5 country_legal:',cl('country_legal',{'1_common law','2_civil law','99_NCE'}))
print('W6 regulation:',cl('regulation_sample_start',{'with ETS/CT','without ETS/CT','99_NCE'}),
      '| q_status:',cl('q_status',{'0_published','1_not published'}))
print('W7 country coverage:',new['country_region'].notna().sum(),new['country_culture'].notna().sum(),
      new['country_legal'].notna().sum(),new['country_econ'].notna().sum(),'(each /2852)')
print('W8 q_status rows:',new['q_status'].notna().sum(),'| q_VHB (legacy semantics):',new['q_VHB'].notna().sum(),
      '| field:',new['field'].notna().sum())
print('W9 lookup q_status:',lknew['q_status'].notna().sum(),'/120 | q_vhb letters:',lknew['q_vhb'].notna().sum(),
      '/120 | jif:',lknew['jif'].notna().sum(),'/120')
