"""
fix_inject.py — rebuild v11.2 with a well-formed formula layer.
Strategy: openpyxl writes the VALUES (as in v11.1 + new q_VHB/field codings);
then an ElementTree pass converts the target cells to formula cells
(<c t="str"><f>FORMULA</f><v>CACHED</v></c>) — well-formed by construction.
Validation: (1) ET parse of both sheets; (2) LibreOffice headless load+recalc on a
throwaway copy, recalculated values compared to expected; (3) X-verifier vs v11.1.
"""
import pandas as pd, numpy as np, shutil, zipfile, os, re
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

NS='http://schemas.openxmlformats.org/spreadsheetml/2006/main'
ET.register_namespace('', NS)

SRC='CER-COD_data_v11_1_candidate.xlsx'; OUT='CER-COD_data_v11_3_candidate.xlsx'
d=pd.read_excel(SRC,'data'); lk=pd.read_excel(SRC,'lookup')
d10=pd.read_excel('/mnt/project/CERCOD_data_v10.xlsx','data')
cl10=pd.read_excel('/mnt/project/CERCOD_data_v10.xlsx','country_lookup')
ext=pd.read_csv('country_lookup_extension_v11_1.csv')
EXC={'Bannier et al (2022)','Fard et al (2020)','Nemoto, Liu (2020)','Srivisal et al (2021)'}

# ---- P-03 regulation recode (rule-based; only NaN cells filled; coded cells never overwritten) ----
REG_RULES = {
 # study: (start_fill, end_fill, rule, flag)
 'Azmi et al (2021)':          ('99_NCE','99_NCE','mixed emerging-economy jurisdictions -> NCE',0),
 'Chodnicka-Jaworska (2022)':  ('99_NCE','99_NCE','Worldwide -> NCE',0),
 'Cubas, Martinez (2018)':     ('99_NCE','99_NCE','Worldwide -> NCE',0),
 'Delis et al (2021)':         ('99_NCE','99_NCE','global mixed jurisdictions -> NCE',0),
 'Ferriani (2022)':            ('99_NCE','99_NCE','Worldwide -> NCE',0),
 'Hoepner et al (2016)':       ('99_NCE','99_NCE','global mixed jurisdictions -> NCE',0),
 'Salvi et al (2021)':         ('99_NCE','99_NCE','Worldwide -> NCE',0),
 'Temiz (2022)':               ('99_NCE','99_NCE','global mixed jurisdictions -> NCE',0),
 'Wang et al (2020)':          ('99_NCE','99_NCE','global mixed jurisdictions -> NCE',0),
 'Yilmaz (2022)':              ('99_NCE','99_NCE','mixed incl. AUS (CT repealed 2014) / CAN -> heterogeneous -> NCE',1),
 'Höck et al (2020)':          ('with ETS/CT','with ETS/CT','all-European sample; EU ETS since 2005; per-row windows unreported',1),
 'Kordschia (2020)':           ('with ETS/CT','with ETS/CT','euro-area list, window 2014-2017; EU ETS homogeneous',0),
 'Christ et al (2022)':        (None,'with ETS/CT','start coded with ETS/CT; EU-only sample, no repeal -> end=with (accretion)',0),
 'Drago et al (2018)':         (None,'with ETS/CT','start coded with ETS/CT; EU-only sample -> end=with (accretion)',0),
 'Sze et al (2021)':           ('without ETS/CT',None,'end coded without; start mirrored (accretion; AUS repeal caveat)',1),
}
reg_log=[]
for stx,(fs,fe,rule,flag) in REG_RULES.items():
    m=d['study']==stx
    n1=n2=0
    if fs is not None:
        t=m & d['regulation_sample_start'].isna(); d.loc[t,'regulation_sample_start']=fs; n1=int(t.sum())
    if fe is not None:
        t=m & d['regulation_sample_end'].isna(); d.loc[t,'regulation_sample_end']=fe; n2=int(t.sum())
    reg_log.append(dict(study=stx,start_filled=n1,end_filled=n2,start_value=fs or '(coded)',
                        end_value=fe or '(coded)',rule=rule,flag='FLAG' if flag else ''))
REG_TOUCHED=True

# ---- country_map ----
rows=[]
for _,r in cl10.iterrows():
    rows.append(dict(country_str=str(r['country']),region=r['region'],econ=r['development'],
                     culture=r['culture'],legal=r['law'],source='v10_country_lookup',note=str(r.get('note',''))))
E=ext.to_dict('records')
def classify(s):
    h=[e for e in E if str(s).startswith(e['mod_country_key'])] or [e for e in E if e['mod_country_key'] in str(s)]
    return max(h,key=lambda e:len(e['mod_country_key'])) if h else None
econ_obs=d.groupby('study')['country_econ'].first()
seen={r['country_str'] for r in rows}
for st,ms in lk.set_index('study')['mod_country'].dropna().items():
    if ms in seen: continue
    h=classify(ms); assert h, ms
    rows.append(dict(country_str=ms,region=h['region'],econ=econ_obs[st],culture=h['culture'],
                     legal=h['legal'],source='extension_v11_1',note=h['rule']+(' [FLAG]' if h.get('flag') else '')))
    seen.add(ms)
cmap=pd.DataFrame(rows); cmap.insert(0,'ckey',[f'CM-{i:03d}' for i in range(1,len(cmap)+1)])
ck=cmap.set_index('country_str')['ckey']

# ---- lookup values ----
cstr10=d10.groupby('study')['country'].first()
lk['country_str']=lk['study'].map(cstr10)
mm=lk['country_str'].isna()
lk.loc[mm,'country_str']=lk.loc[mm,'study'].map(lk.set_index('study')['mod_country'])
lk.loc[lk['study'].isin(EXC),'country_str']='row-level (see data; within-study variation)'
lk['ckey']=lk['country_str'].map(ck)
fmap=d10.groupby('study')['field'].first()
lk['field']=lk['study'].map(fmap)
FJ={'business strategy and the environment':'2_sust','corporate social responsibility':'2_sust',
    'environmental management':'2_sust','cleaner production':'2_sust','sustainability':'2_sust',
    'environmental science and pollution':'2_sust','energies':'2_sust',
    'corporate governance':'3_mgmt','sage open':'3_mgmt','ethics and systems':'3_mgmt'}
FLAGJ={'energies','sage open','ethics and systems','european research studies','economics and business review'}
def fieldcode(j):
    if pd.isna(j) or str(j).strip().lower() in {'nan',''}: return '1_fin/acc/econ','WP, content-based [FLAG]'
    s=str(j).lower()
    for k,v in FJ.items():
        if k in s: return v,'keyword: '+k+(' [FLAG]' if any(f in s for f in FLAGJ) else '')
    return '1_fin/acc/econ','finance/econ default'+(' [FLAG]' if any(f in s for f in FLAGJ) else '')
notes=[]
for i,r in lk.iterrows():
    if pd.isna(r['field']):
        v,n=fieldcode(r['journal_full']); lk.at[i,'field']=v; notes.append(n)
    else: notes.append('v10 carry')
lk['field_note']=notes
def q_status_class(s): return '1_not published' if s=='working paper' else '0_published'
def q_vhb_class(s,l):
    # P-09 final rule (author 2026-07-10): WP -> 99_NCE; published A+/A/B/C -> high;
    # published D or not VHB-ranked -> low
    if s=='working paper': return '99_NCE'
    if pd.isna(l) or str(l).strip()=='': return '0_VHB low'
    return '1_VHB high' if l in {'A+','A','B','C'} else '0_VHB low'
lk['q_status_class']=[q_status_class(s) for s in lk['q_status']]
lk['q_vhb_class']=[q_vhb_class(s,l) for s,l in zip(lk['q_status'],lk['q_vhb'])]
cmI=cmap.set_index('ckey')
for c,mc in [('c_region','region'),('c_econ','econ'),('c_culture','culture'),('c_legal','legal')]:
    lk[c]=lk['ckey'].map(cmI[mc])
lkI=lk.set_index('study')
exp=pd.DataFrame({'study':d['study']})
exp['q_status']=exp['study'].map(lkI['q_status_class'])
exp['q_VHB']=exp['study'].map(lkI['q_vhb_class'])
exp['field']=exp['study'].map(lkI['field'])
for c,lc in [('country_region','c_region'),('country_econ','c_econ'),
             ('country_culture','c_culture'),('country_legal','c_legal')]:
    exp[c]=exp['study'].map(lkI[lc])
# exception rows keep v11.1 hardcoded country values
for c in ['country_region','country_econ','country_culture','country_legal']:
    m=d['study'].isin(EXC); exp.loc[m,c]=d.loc[m,c]

# ---- write VALUES via openpyxl ----
shutil.copy(SRC,OUT); wb=load_workbook(OUT)
if 'country_map' in wb.sheetnames: del wb['country_map']
wm=wb.create_sheet('country_map')
for j,c in enumerate(cmap.columns,1): wm.cell(1,j).value=c
for i,(_,r) in enumerate(cmap.iterrows(),2):
    for j,c in enumerate(cmap.columns,1): wm.cell(i,j).value=None if pd.isna(r[c]) else r[c]
wl=wb['lookup']; base=wl.max_column
newcols=['country_str','ckey','field','field_note','q_status_class','q_vhb_class','c_region','c_econ','c_culture','c_legal']
pos={c:base+k+1 for k,c in enumerate(newcols)}; L={c:get_column_letter(j) for c,j in pos.items()}
for c,j in pos.items(): wl.cell(1,j).value=c
lk_f={}
for i in range(2,len(lk)+2):
    r=lk.iloc[i-2]
    for c in newcols:
        wl.cell(i,pos[c]).value=None if pd.isna(r[c]) else r[c]
    lk_f[f'{L["q_status_class"]}{i}']=f'IF($P{i}="working paper","1_not published","0_published")'
    lk_f[f'{L["q_vhb_class"]}{i}']=(f'IF($P{i}="working paper","99_NCE",'
                                    f'IF(OR($Q{i}="A+",$Q{i}="A",$Q{i}="B",$Q{i}="C"),"1_VHB high","0_VHB low"))')
    if pd.notna(r['ckey']):
        for c,mapcol in [('c_region','C'),('c_econ','D'),('c_culture','E'),('c_legal','F')]:
            lk_f[f'{L[c]}{i}']=(f'IFERROR(INDEX(country_map!${mapcol}:${mapcol},'
                                f'MATCH(${L["ckey"]}{i},country_map!$A:$A,0)),"")')
ws=wb['data']; hdr={ws.cell(1,j).value:j for j in range(1,ws.max_column+1)}
# write recoded regulation values (hardcoded coding, not formula-derived)
for col in ['regulation_sample_start','regulation_sample_end']:
    j=hdr[col]
    for i,v in enumerate(d[col].tolist(),start=2):
        ws.cell(i,j).value=None if pd.isna(v) else v
# reg recode log sheet
if 'reg_recode_log' in wb.sheetnames: del wb['reg_recode_log']
wr=wb.create_sheet('reg_recode_log')
import pandas as _pd
_rl=_pd.DataFrame(reg_log)
for j,c in enumerate(_rl.columns,1): wr.cell(1,j).value=c
for i,(_,r) in enumerate(_rl.iterrows(),2):
    for j,c in enumerate(_rl.columns,1): wr.cell(i,j).value=r[c]
FMAP={'q_status':L['q_status_class'],'q_VHB':L['q_vhb_class'],'field':L['field'],
      'country_region':L['c_region'],'country_econ':L['c_econ'],
      'country_culture':L['c_culture'],'country_legal':L['c_legal']}
CCOLS={'country_region','country_econ','country_culture','country_legal'}
jstudy=hdr['study']; d_f={}
for i in range(2,len(d)+2):
    st=ws.cell(i,jstudy).value
    for col,lcol in FMAP.items():
        j=hdr[col]
        ws.cell(i,j).value=None if pd.isna(exp.iloc[i-2][col]) else exp.iloc[i-2][col]
        if not (col in CCOLS and st in EXC):
            d_f[f'{get_column_letter(j)}{i}']=f'INDEX(lookup!${lcol}:${lcol},MATCH($A{i},lookup!$A:$A,0))'
wb.save(OUT)

# ---- ET pass: convert target cells to formula cells ----
def add_formulas(path, sheetname, fmap_cells, valmap):
    with zipfile.ZipFile(path) as z:
        wbxml=z.read('xl/workbook.xml').decode('utf8')
        rels=z.read('xl/_rels/workbook.xml.rels').decode('utf8')
        rid=re.search(r'<sheet[^>]*name="%s"[^>]*r:id="(rId\d+)"'%sheetname, wbxml).group(1)
        tgt=re.search(r'<Relationship\b(?=[^>]*Id="%s")[^>]*Target="([^"]+)"'%rid, rels).group(1)
        tgt=tgt.lstrip('/');  tgt=tgt if tgt.startswith('xl/') else 'xl/'+tgt
        contents={n:z.read(n) for n in z.namelist()}
    root=ET.fromstring(contents[tgt])
    sd=root.find(f'{{{NS}}}sheetData')
    n=0
    for row in sd.findall(f'{{{NS}}}row'):
        for c in row.findall(f'{{{NS}}}c'):
            ref=c.get('r')
            if ref in fmap_cells:
                for ch in list(c): c.remove(ch)
                c.set('t','str')
                f=ET.SubElement(c,f'{{{NS}}}f'); f.text='='+fmap_cells[ref] if False else fmap_cells[ref]
                v=ET.SubElement(c,f'{{{NS}}}v'); v.text=str(valmap.get(ref,''))
                n+=1
    contents[tgt]=b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'+ET.tostring(root)
    tmp=path+'.tmp'
    with zipfile.ZipFile(tmp,'w',zipfile.ZIP_DEFLATED) as z:
        for name,b in contents.items(): z.writestr(name,b)
    os.replace(tmp,path)
    return n

# cached values per cell (what openpyxl just wrote)
dval={}
for i in range(2,len(d)+2):
    for col in FMAP:
        j=hdr[col]; ref=f'{get_column_letter(j)}{i}'
        if ref in d_f:
            v=exp.iloc[i-2][col]; dval[ref]='' if pd.isna(v) else v
lval={}
for i in range(2,len(lk)+2):
    r=lk.iloc[i-2]
    for c in ['q_status_class','q_vhb_class','c_region','c_econ','c_culture','c_legal']:
        ref=f'{L[c]}{i}'
        if ref in lk_f:
            v=r[c]; lval[ref]='' if pd.isna(v) else v
n1=add_formulas(OUT,'data',d_f,dval)
n2=add_formulas(OUT,'lookup',lk_f,lval)
print('formula cells written: data',n1,'lookup',n2)
